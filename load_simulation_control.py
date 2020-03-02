from __future__ import print_function
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except:
    print("Openpyxl module cannot be imported because it's not installed. pip install openpyxl to solve the issue")
    exit('Simulation terminated')
from re import compile, escape
from scipy import interpolate
from fipy import numerix
from collections import OrderedDict

# User Setup
simulation_control_workbook = 'sim_settings.xlsx'
simulation_control_workbook_sheet = 'Sim_Control'

################################# Import the simuatlion settings #######################################################
cell_properties  = load_workbook(simulation_control_workbook)       # Load workbook containing cell property data
sim_settings_sheet = cell_properties.get_sheet_by_name(simulation_control_workbook_sheet)  # Load the sheet corresponding to desired cell
Uocp_neg_sheet = 'Uocp_neg'
Uocp_pos_sheet = 'Uocp_pos'
kappa_sheet = 'Kappa'

control_chars = ''.join(map(unichr, range(0,32) + range(127,160)))           # Deal with non-printable characters in cell property names
control_char_re = compile('[%s]' % escape(control_chars))
def remove_control_chars(s):
    return control_char_re.sub('', s.strip())

for sheet in cell_properties:
    if sheet.title == sim_settings_sheet.title:
        namecol = sim_settings_sheet.min_column  # This column no. contains the name of the cell parameters
        minrow = sim_settings_sheet.min_row  # This column no. contains the name of the cell parameters
        valuecol = namecol + 1  # This column no. contains the numerical/string value of the cell-parameters
        # settings_dict = {}  # Create an object to store the parameters
        settings_dict = OrderedDict()
        for row_number in range(minrow, sim_settings_sheet.max_row+1):                             # Import cell properties sim_settings_sheet
            name_location  = get_column_letter(namecol)  + str(row_number)          # Obtain the alphabetical column string of the name column
            value_location = get_column_letter(valuecol) + str(row_number)          # Obtain the alphabetical column string of the value column
            if sim_settings_sheet[name_location].value != None:                                    # Skip processing if blank row
                sim_setting_name = remove_control_chars(sim_settings_sheet[name_location].value)   # Get the contents of the 'name-column' & strip any non-printable characters
                sim_setting_name = sim_settings_sheet[name_location].value
                sim_setting_value = sim_settings_sheet[value_location].value                       # Force to type 'float' if user entered a cell parameter without decimal point
                if not isinstance(sim_settings_sheet[value_location].value, float):
                    if isinstance(sim_settings_sheet[value_location].value, long or int):
                        sim_setting_value = float(sim_setting_value)
                    else:
                        sim_setting_value = str(sim_setting_value)             # Convert unicode Uocp functions to strings
                settings_dict[sim_setting_name] = sim_setting_value

    elif sheet.title == Uocp_neg_sheet:
        namecol = sheet.min_column  # This column no. contains the timestamp
        valuecol = namecol + 1  # This column no. contains the current value
        minrow = sheet.min_row
        maxrow = sheet.max_row
        data_col_length = maxrow - minrow + 1
        theta_table_vals_neg = numerix.empty(data_col_length)
        Uocp_table_vals_neg = numerix.empty(data_col_length)
        for row_number in range(minrow, maxrow + 1):
            name_location = get_column_letter(namecol) + str(row_number)  # Obtain the alphabetical column string of the timestamp column
            value_location = get_column_letter(valuecol) + str(row_number)  # Obtain the alphabetical column string of the current_value column
            if sheet[name_location].value != None:  # Check for blank row
                theta_table_vals_neg[row_number - minrow] = float(sheet[name_location].value)
                Uocp_table_vals_neg[row_number - minrow] = float(sheet[value_location].value)  # Need to add in ability to process formulae in spreadsheet cells

    elif sheet.title == Uocp_pos_sheet:
        namecol = sheet.min_column  # This column no. contains the timestamp
        valuecol = namecol + 1  # This column no. contains the current value
        minrow = sheet.min_row
        maxrow = sheet.max_row
        data_col_length = maxrow - minrow + 1
        theta_table_vals_pos = numerix.empty(data_col_length)
        Uocp_table_vals_pos = numerix.empty(data_col_length)
        for row_number in range(minrow, maxrow + 1):
            name_location = get_column_letter(namecol) + str(row_number)  # Obtain the alphabetical column string of the timestamp column
            value_location = get_column_letter(valuecol) + str(row_number)  # Obtain the alphabetical column string of the current_value column
            if sheet[name_location].value != None:  # Check for blank row
                theta_table_vals_pos[row_number - minrow] = float(sheet[name_location].value)
                Uocp_table_vals_pos[row_number - minrow] = float(sheet[value_location].value)  # Need to add in ability to process formulae in spreadsheet cells

    elif sheet.title == kappa_sheet:
        namecol = sheet.min_column  # This column no. contains the timestamp
        valuecol = namecol + 1  # This column no. contains the current value
        minrow = sheet.min_row
        maxrow = sheet.max_row
        data_col_length = maxrow - minrow + 1
        kappa_ce = numerix.empty(data_col_length)
        kappa_value = numerix.empty(data_col_length)
        for row_number in range(minrow, maxrow + 1):
            name_location = get_column_letter(namecol) + str(row_number)  # Obtain the alphabetical column string of the timestamp column
            value_location = get_column_letter(valuecol) + str(row_number)  # Obtain the alphabetical column string of the current_value column
            if sheet[name_location].value != None:  # Check for blank row
                kappa_ce[row_number - minrow] = float(sheet[name_location].value)
                kappa_value[row_number - minrow] = float(sheet[value_location].value)  # Need to add in ability to process formulae in spreadsheet cells
        interpolated_kappa_fn = interpolate.interp1d(kappa_ce, kappa_value, kind='slinear')

#################################################### Cleanse names #####################################################
# General
initial_soc_percent, cell_temperature = settings_dict['initial_soc_percent'], settings_dict['cell_temperature']

# Timing
simEndtime, temporal_resolution = settings_dict['simEndtime'], settings_dict['temporal_resolution']

# Domain sizing
L_neg_normalised, L_pos_normalised = settings_dict['L_neg_normalised'], settings_dict['L_pos_normalised']
L_sep_normalised = settings_dict['L_sep_normalised']
Rs_neg_normalised, Rs_pos_normalised = settings_dict['Rs_neg_normalised'], settings_dict['Rs_pos_normalised']

# Domain meshing
nx_neg, nx_pos, nx_sep = int(settings_dict['nx_neg']), int(settings_dict['nx_pos']), int(settings_dict['nx_sep'])
nr_neg, nr_pos = int(settings_dict['nr_neg']), int(settings_dict['nr_pos'])
non_uniform_axial_meshing, non_uniform_radial_meshing = settings_dict['non_uniform_axial_meshing'], settings_dict['non_uniform_radial_meshing']

# Solution variable tolerancing
tolerance_phi_s_neg_res, tolerance_phi_s_pos_res = settings_dict['tolerance_phi_s_neg_res'], settings_dict['tolerance_phi_s_pos_res']
tolerance_Cs_neg_res, tolerance_Cs_pos_res = settings_dict['tolerance_Cs_neg_res'], settings_dict['tolerance_Cs_pos_res']
tolerance_phi_e_sim_cell_res, tolerance_Ce_sim_cell_res = settings_dict['tolerance_phi_e_sim_cell_res'], settings_dict['tolerance_Ce_sim_cell_res']

# Under-relaxation factors
under_relax_factor_phi_s_neg = settings_dict['under_relax_factor_phi_s_neg']
under_relax_factor_phi_s_pos = settings_dict['under_relax_factor_phi_s_pos']
under_relax_factor_Ce = settings_dict['under_relax_factor_Ce']
under_relax_factor_phi_e = settings_dict['under_relax_factor_phi_e']
under_relax_factor_Cs_neg = settings_dict['under_relax_factor_Cs_neg']
under_relax_factor_Cs_pos = settings_dict['under_relax_factor_Cs_pos']

# Current profile import
cprofile_workbook, cprofile_sheet, repeat_current_profile = simulation_control_workbook, settings_dict['cprofile_sheet'], settings_dict['repeat_current_profile']

# Cell properties import
cell_properties_workbook, cell_properties_sheet = simulation_control_workbook, settings_dict['cell_properties_sheet']

# Data logging & Debugging
datalogging_switch = settings_dict['datalogging_switch']
if datalogging_switch == 'on':
    datalogging_switch == 'On'
record_period, datalog_filename, output_location = settings_dict['record_period'], settings_dict['datalog_filename'], settings_dict['output_location']
det_residual_live_plot, det_residual_saving = settings_dict['det_residual_live_plot'], settings_dict['det_residual_saving']
mem_leak_analysis, det_concentrations_live_plot = settings_dict['mem_leak_analysis'], settings_dict['Cs_live_plot']
if det_residual_live_plot == 'on':
    det_residual_live_plot == 'On'
if det_residual_saving == 'on':
    det_residual_live_plot == 'On'
if mem_leak_analysis == 'on':
    mem_leak_analysis == 'On'
if det_concentrations_live_plot == 'on':
    det_concentrations_live_plot == 'On'

# Analysis
auto_process = settings_dict['auto_process']
if auto_process == 'on':
    auto_process == 'On'
if auto_process == 'On' and datalogging_switch != 'On':
    print('auto_process data is switched On, but datalogging is switched off. Either switch datalogging to On or switch auto_process to Off')
    exit('Simulation terminated')

########################################### Pre-process domain sizes ###################################################
nx_tot = nx_neg + nx_sep + nx_pos
L_tot_normalised = L_neg_normalised + L_pos_normalised + L_sep_normalised  # Total normalised length (by default set to 3)


########################### Current Profile Import #####################################################################
cprofile_book = load_workbook(cprofile_workbook)                                   # Load workbook containing current profile
cprofile_data = cprofile_book.get_sheet_by_name(cprofile_sheet)                             # Load the sheet corresponding to desired profile

cprofile_namecol = cprofile_data.min_column                                                 # This column no. contains the timestamp
cprofile_valuecol = cprofile_namecol + 1                                                    # This column no. contains the current value
# Improve the below logic to handle column headers, keep moving down until you hit the first float
cprofile_minrow = cprofile_data.min_row
cprofile_maxrow = cprofile_data.max_row
cprofile_length = cprofile_maxrow - cprofile_minrow + 1
cprofile_timestamps = numerix.empty(cprofile_length)
cprofile_currents = numerix.empty(cprofile_length)

for row_number in range(cprofile_minrow,cprofile_maxrow + 1):
    cprofile_timestamp_location = get_column_letter(cprofile_namecol) + str(row_number)    # Obtain the alphabetical column string of the timestamp column
    cprofile_value_location = get_column_letter(cprofile_valuecol) + str(row_number)       # Obtain the alphabetical column string of the current_value column
    if cprofile_data[cprofile_timestamp_location].value != None:                            # Check for blank row
        cprofile_timestamps[row_number - cprofile_minrow] = float(cprofile_data[cprofile_timestamp_location].value)
        cprofile_currents[row_number - cprofile_minrow] = float(cprofile_data[cprofile_value_location].value) # Need to add in ability to process formulae in spreadsheet cells

interpolated_current_fn = interpolate.interp1d(cprofile_timestamps, cprofile_currents, kind = 'slinear')


################################# Import the cell properties ###########################################################
cell_properties  = load_workbook(cell_properties_workbook)          # Load workbook containing cell property data
data             = cell_properties.get_sheet_by_name(cell_properties_sheet)  # Load the sheet corresponding to desired cell
namecol          = data.min_column                                           # This column no. contains the name of the cell parameters
minrow           = data.min_row                                              # This column no. contains the name of the cell parameters
valuecol         = namecol + 1                                               # This column no. contains the numerical/string value of the cell-parameters
params_dict = {}                                                             # Create an object to store the parameters

for row_number in range(minrow, data.max_row+1):                             # Import cell properties data
    name_location  = get_column_letter(namecol)  + str(row_number)          # Obtain the alphabetical column string of the name column
    value_location = get_column_letter(valuecol) + str(row_number)          # Obtain the alphabetical column string of the value column
    if data[name_location].value != None:                                    # Skip processing if blank row
        cell_parameter_name = remove_control_chars(data[name_location].value)# Get the contents of the 'name-column' & strip any non-printable characters
        cell_parameter_name = data[name_location].value
        cell_parameter_value = data[value_location].value                    # Force to type 'float' if user entered a cell parameter without decimal point
        if not isinstance(data[value_location].value, float):
            if isinstance(data[value_location].value, long or int):
                cell_parameter_value = float(cell_parameter_value)
            else:
                cell_parameter_value = str(cell_parameter_value)             # Convert unicode Uocp functions to strings
        params_dict[cell_parameter_name] = cell_parameter_value
